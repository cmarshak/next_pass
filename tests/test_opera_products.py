from __future__ import annotations

import json
import zipfile
from datetime import UTC, datetime
from unittest.mock import MagicMock, patch

import pytest
from tests.helpers import FakeFrame, FakePolygon

import next_pass.opera_products as opera_products


def test_describe_cloud_cover_thresholds():
    assert "mostly clear" in opera_products.describe_cloud_cover(25)
    assert "partly cloudy" in opera_products.describe_cloud_cover(60)
    assert "mostly cloudy" in opera_products.describe_cloud_cover(90)


def test_find_print_available_opera_products_validates_date_range(tmp_path):
    with pytest.raises(ValueError):
        opera_products.find_print_available_opera_products(
            ["34.2", "-118.17"], 2, "2026-03-10/2026-03-01", None, tmp_path
        )


def test_find_print_available_opera_products_prefixes_products_and_trims_dates(
    monkeypatch, tmp_path
):
    searches = []
    rows = FakeFrame(
        [
            {
                "BeginningDateTime": datetime(2026, 3, 20, 10, tzinfo=UTC),
                "geometry": FakePolygon("g1"),
            },
            {
                "BeginningDateTime": datetime(2026, 3, 20, 12, tzinfo=UTC),
                "geometry": FakePolygon("g2"),
            },
            {
                "BeginningDateTime": datetime(2026, 3, 18, 8, tzinfo=UTC),
                "geometry": FakePolygon("g3"),
            },
        ]
    )

    def fake_search(short_name, cloud_hosted, bounding_box, temporal, return_gdf):
        searches.append(short_name)
        return (
            [{"id": "a"}, {"id": "b"}, {"id": "c"}],
            rows.copy(),
        )

    monkeypatch.setattr(opera_products.leafmap, "nasa_data_search", fake_search)
    monkeypatch.setattr(opera_products, "bbox_type", lambda bbox: bbox)
    monkeypatch.setattr(
        opera_products,
        "bbox_to_geometry",
        lambda bbox, timestamp_dir: (FakePolygon("aoi"), [0, 1, 2, 3], None),
    )
    monkeypatch.setattr(opera_products.pd, "to_datetime", lambda values: values)
    monkeypatch.setattr(opera_products.time, "sleep", lambda seconds: None)

    result = opera_products.find_print_available_opera_products(
        bbox=[34.2, -118.17],
        number_of_dates=1,
        date_str="2026-03-23",
        list_of_products=["RTC-S1_V1", "DSWX-HLS_V1"],
        timestamp_dir=tmp_path,
    )

    assert searches == ["OPERA_L2_RTC-S1_V1", "OPERA_L3_DSWX-HLS_V1"]
    assert len(result["OPERA_L2_RTC-S1_V1"]["results"]) == 2


def test_export_opera_products_writes_workbook_and_skips_cloudiness_when_disabled(
    tmp_path,
):
    geometry = FakePolygon("geom")
    results_dict = {
        "OPERA_L3_DSWX-HLS_V1": {
            "results": [
                {
                    "umm": {
                        "GranuleUR": "granule-1",
                        "TemporalExtent": {
                            "RangeDateTime": {
                                "BeginningDateTime": "2026-03-20T00:00:00Z",
                                "EndingDateTime": "2026-03-20T01:00:00Z",
                            }
                        },
                        "RelatedUrls": [
                            {"URL": "https://example.com/file_B01_WTR.tif"},
                            {"URL": "https://example.com/file_CLOUD.tif"},
                        ],
                    }
                }
            ],
            "gdf": FakeFrame([{"geometry": geometry}]),
        }
    }

    opera_products.export_opera_products(
        results_dict,
        tmp_path,
        compute_cloudiness=False,
    )

    output_file = tmp_path / "opera_products_metadata.xlsx"
    assert output_file.exists()
    if zipfile.is_zipfile(output_file):
        from openpyxl import load_workbook

        workbook = load_workbook(output_file)
        sheet = workbook["OPERA Metadata"]
        assert sheet["A1"].value == "Dataset"
        assert sheet["B2"].value == "granule-1"
        assert sheet["F2"].value == "https://example.com/file_B01_WTR.tif"
    else:
        payload = json.loads(output_file.read_text(encoding="utf-8"))
        assert payload[0][0] == "Dataset"
        assert payload[1][1] == "granule-1"
        assert payload[1][5] == "https://example.com/file_B01_WTR.tif"


@patch("next_pass.opera_products.earthaccess.search_data")
def test_fetch_hls_granule_links(mock_search):
    # Setup mock successful response
    mock_result = MagicMock()
    mock_result.data_links.return_value = ["https://example.com/B04.tif"]
    mock_search.return_value = [mock_result]

    # Test 1: Successful fetch
    links = opera_products.fetch_hls_granule_links("HLS.S30.T11SLT.20230101.v2.0")
    assert links == ["https://example.com/B04.tif"]
    mock_search.assert_called_with(
        short_name="HLSS30", granule_name="HLS.S30.T11SLT.20230101.v2.0"
    )

    # Test 2: Error case (API failure)
    mock_search.side_effect = Exception("API Timeout")
    error_links = opera_products.fetch_hls_granule_links("HLS.S30.T11SLT.20230101.v2.0")
    assert error_links is None


@patch("next_pass.opera_products.fetch_hls_granule_links")
def test_export_hls_band_mapping_and_input_granules(mock_fetch_links, tmp_path):
    # Simulate an S30 and L30 response with their specific band names
    mock_fetch_links.side_effect = [
        [
            "https://fake/B02.tif",
            "https://fake/B03.tif",
            "https://fake/B04.tif",
            "https://fake/B8A.tif",
            "https://fake/Fmask.tif",
        ],  # S30 response
        [
            "https://fake/B02.tif",
            "https://fake/B03.tif",
            "https://fake/B04.tif",
            "https://fake/B05.tif",
            "https://fake/Fmask.tif",
        ],  # L30 response
    ]

    mock_results = {
        "OPERA_L3_DSWX-HLS_V1": {
            "results": [
                {
                    "umm": {
                        "GranuleUR": "OPERA_S30",
                        "InputGranules": ["HLS.S30.T11SLT"],
                    }
                },
                {
                    "umm": {
                        "GranuleUR": "OPERA_L30",
                        "InputGranules": ["HLS.L30.T11SLT"],
                    }
                },
            ],
            "gdf": None,  # Skipping geometry for simplicity
        }
    }

    # Run the export function
    opera_products.export_opera_products(
        mock_results, tmp_path, compute_cloudiness=False, include_hls=True
    )

    # Read the generated Excel file to verify mappings
    from openpyxl import load_workbook

    wb = load_workbook(tmp_path / "opera_products_metadata.xlsx")
    ws = wb.active

    # Row 2 is S30 (B8A for NIR)
    assert ws.cell(row=2, column=18).value == "https://fake/B04.tif"  # Red
    assert ws.cell(row=2, column=21).value == "https://fake/B8A.tif"  # NIR
    assert ws.cell(row=2, column=22).value == "https://fake/Fmask.tif"

    # Row 3 is L30 (B05 for NIR)
    assert ws.cell(row=3, column=21).value == "https://fake/B05.tif"  # NIR for Landsat


@patch("next_pass.opera_products.earthaccess.search_data")
def test_export_hls_fallback_cmr_search(mock_cmr_search, tmp_path):
    # Setup mock geometry and mock CMR response
    mock_geom = MagicMock()
    mock_geom.bounds = (-120, 30, -119, 31)
    mock_geom.wkt = "POLYGON((-120 30, -119 30, -119 31, -120 31, -120 30))"

    mock_hls_result = MagicMock()
    mock_hls_result.get.return_value = {"GranuleUR": "HLS.S30.T11SLT.123"}
    mock_hls_result.data_links.return_value = ["https://fallback/B04.tif"]
    mock_cmr_search.return_value = [mock_hls_result]

    # Provide results missing "InputGranules", forcing the fallback logic
    mock_results = {
        "OPERA_L3_DIST-ALERT-HLS_V1": {
            "results": [
                {
                    "umm": {
                        "GranuleUR": "OPERA_L3_DIST-ALERT-HLS_T11SLT_20230101",
                        "TemporalExtent": {
                            "RangeDateTime": {
                                "BeginningDateTime": "2023-01-01T00:00:00Z"
                            }
                        },
                    }
                }
            ],
            "gdf": FakeFrame([{"geometry": mock_geom}]),
        }
    }

    opera_products.export_opera_products(
        mock_results, tmp_path, compute_cloudiness=False, include_hls=True
    )

    # Verify fallback search triggered with bounds and date
    mock_cmr_search.assert_called_once()
    kwargs = mock_cmr_search.call_args.kwargs
    assert kwargs["bounding_box"] == (-120, 30, -119, 31)
    assert kwargs["temporal"] == ("2023-01-01T00:00:00", "2023-01-01T23:59:59")

    # Verify the fallback successfully mapped the data
    from openpyxl import load_workbook

    wb = load_workbook(tmp_path / "opera_products_metadata.xlsx")
    ws = wb.active
    assert (
        ws.cell(row=2, column=17).value == "HLS.S30.T11SLT.123"
    )  # Extracted Fallback ID
    assert ws.cell(row=2, column=18).value == "https://fallback/B04.tif"
